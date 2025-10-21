import json
import boto3
from boto3.dynamodb.conditions import Key, Attr
from datetime import datetime, timedelta
import os
import csv
import pandas as pd
from aws_lambda_powertools import Logger, Tracer
from custom_exceptions import BadRequestException, ResourceNotFoundException
from zipfile import ZipFile
import re
from dateutil import parser
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
import requests
from requests_aws4auth import AWS4Auth

SMART_EYE_BUCKET = os.environ.get('SMART_EYE_BUCKET')  
DOWNLOAD_JOB_TABLE = os.environ.get('DOWNLOAD_JOB_TABLE')
FIXED_ASSET_TABLE = os.environ.get('FIXED_ASSET_TABLE')
ACQUISITION_JOURNAL_TABLE = os.environ.get('ACQUISITION_JOURNAL_TABLE')
MERCHANT_TABLE = os.environ.get('MERCHANT_TABLE') 
ES_DOMAIN_ENDPOINT = os.environ.get('ES_DOMAIN_ENDPOINT')

# AWS Clients and Resources
S3_CLIENT = boto3.client('s3')
DDB_RESOURCE = boto3.resource('dynamodb')

# DynamoDB Tables
DOWNLOAD_JOB_DDB_TABLE = DDB_RESOURCE.Table(DOWNLOAD_JOB_TABLE) if DOWNLOAD_JOB_TABLE else None
FIXED_ASSET_DDB_TABLE = DDB_RESOURCE.Table(FIXED_ASSET_TABLE)
ACQUISITION_JOURNAL_DDB_TABLE = DDB_RESOURCE.Table(ACQUISITION_JOURNAL_TABLE)
MERCHANT_DDB_TABLE = DDB_RESOURCE.Table(MERCHANT_TABLE) 

ZERO_WIDTH_SPACE = '\u200B'  

logger = Logger()
tracer = Tracer()

CREDENTIALS = boto3.Session().get_credentials()
ACCESS_KEY = CREDENTIALS.access_key
SECRET_ACCESS_KEY = CREDENTIALS.secret_key
AWSAUTH = AWS4Auth(ACCESS_KEY, SECRET_ACCESS_KEY, 'ap-southeast-1', 'es', session_token=CREDENTIALS.token)

"""
SAMPLE PAYLOAD
{
    "filter": {
        "and": []
    },
    "module": "FixedAsset",
    "resultType": "fixedAsset",  // or "acquisitionJournal" or "all"
    "sort": {
        "direction": "desc",
        "field": "createdAt"
    },
    "merchantId": "default",
    "jobId": "job-123",
    "outputType": "csv"
}
"""

@tracer.capture_lambda_handler
def lambda_handler(event, context):
    try:
        arguments = event.get('arguments', {})
        merchant_id = event.get('merchantId', 'default')
        job_id = event.get('jobId')
        result_type = arguments.get('resultType', 'all')
        output_type = arguments.get('outputType', 'csv')
        item_id_list = arguments.get('itemIdList', [])

        logger.info(f"Processing export for merchant_id: {merchant_id}, jobId: {job_id}, resultType: {result_type}, output_type: {output_type}")

        # ADDED: Get merchant configuration (following ExportGeneratedPOFINAL pattern)
        merchant_config = getMerchantConfiguration(merchant_id)
        
        # ADDED: Check if merchant wants Excel format and override output_type
        custom_logics = merchant_config.get('customLogics', {})
        exportExcelFormat = custom_logics.get('exportExcelFormat', False)
        if exportExcelFormat:
            output_type = 'xlsx'  # Override to use xlsx mapping
            file_extension = 'xlsx'
        else:
            file_extension = 'csv'

        # Sort configuration
        sort_field = 'createdAt'
        sort_direction = 'desc'
        if arguments.get('sort') is not None:
            sort_field = arguments.get('sort').get('field', 'createdAt')
            sort_direction = arguments.get('sort').get('direction', 'desc')

        filters = arguments.get('filter', {})
        
        # Process based on result type
        export_results = []
        
        if result_type in ['fixedAsset', 'all']:
            fixed_asset_result = exportFixedAssets(
                merchant_id, filters, sort_field, sort_direction, 
                item_id_list, job_id, output_type, exportExcelFormat, file_extension
            )
            if fixed_asset_result:
                export_results.append(fixed_asset_result)
        
        if result_type in ['acquisitionJournal', 'all']:
            journal_result = exportAcquisitionJournal(
                merchant_id, filters, sort_field, sort_direction,
                item_id_list, job_id, output_type, exportExcelFormat, file_extension
            )
            if journal_result:
                export_results.append(journal_result)
        
        if not export_results:
            updateDownloadJobStatus(job_id, 'COMPLETED', 'No data found')
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'status': True,
                    'message': 'No data found to export'
                })
            }
        
        # Create combined zip if multiple exports
        if len(export_results) > 1:
            combined_result = createCombinedExport(export_results, merchant_id)
            updateDownloadJobStatus(
                job_id, 'COMPLETED', 'Export completed successfully',
                combined_result['s3_path'], combined_result['presigned_url']
            )
        else:
            result = export_results[0]
            updateDownloadJobStatus(
                job_id, 'COMPLETED', 'Export completed successfully',
                result['s3_path'], result['presigned_url']
            )
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'status': True,
                'message': 'Export completed successfully',
                'exports': len(export_results)
            })
        }
    
    except (BadRequestException, ResourceNotFoundException) as ex:
        logger.error(f"Business error: {str(ex)}")
        updateDownloadJobStatus(job_id, 'FAILED', str(ex))
        return {
            'statusCode': 400,
            'body': json.dumps({
                'status': False,
                'message': str(ex)
            })
        }
    
    except Exception as ex:
        updateDownloadJobStatus(job_id, 'FAILED', str(ex))
        tracer.put_annotation("lambda_error", "true")
        tracer.put_annotation("lambda_name", context.function_name)
        tracer.put_metadata("event", event)
        tracer.put_metadata("message", str(ex))
        logger.exception({"message": str(ex)})
        return {
            'statusCode': 500,
            'body': json.dumps({
                'status': False,
                'message': "The server encountered an unexpected condition that prevented it from fulfilling your request."
            })
        }

@tracer.capture_method
def exportFixedAssets(merchant_id, filters, sort_field, sort_direction, item_id_list, job_id, output_type, exportExcelFormat, file_extension):
    """Export Fixed Asset data using Elasticsearch"""
    try:

        fixed_assets_es = getDataFromES(
            merchant_id, 
            sort_field, 
            sort_direction, 
            filters, 
            item_id_list,
            'fixedAsset' 
        )
        
        if not fixed_assets_es:
            logger.info("No fixed asset data found")
            return None
        
        # Extract _source from ES results
        fixed_assets = [item['_source'] for item in fixed_assets_es]
        
        # Get mapping configuration (updated path)
        mapping_config = getMappingConfig(merchant_id, 'fixedAsset', output_type)
        
        # Get export configuration
        export_config = mapping_config.get('exportFixedAsset', {})
        
        # Define headers from mapping config
        headers = export_config.get('headers', [
            "No.", "Description", "Responsible Employee", "FA Class Code",
            "FA Subclass Code", "FA Location Code", "Search Description",
            "Acquired", "Blocked", "FA Posting Group", "Outlet Dim Code",
            "Vendor No"
        ])
        
        # Get field mappings
        document_fields = export_config.get('document_fields', {})
        field_formats = export_config.get('header_formats', {})
        date_parsing_config = export_config.get('date_parsing', {})
        
        # Process data using mapping
        rows = processFixedAssetsWithMapping(fixed_assets, export_config)
        
        # Create file
        current_date_time = datetime.strftime((datetime.now() + timedelta(hours=8)), '%Y-%m-%d_%H:%M:%S')
        filename = f'FixedAsset_{current_date_time}.{file_extension}'
        
        # Clean old files
        clean_temp_files([filename])
        
        if exportExcelFormat:
            exportToExcel(f'/tmp/{filename}', headers, rows, export_config)
        else:
            write_csv_file(filename, headers, rows)
        
        zip_filename = f'FixedAsset_{current_date_time}.zip'
        createZipFile(zip_filename, [filename])
        
        s3_key = f'export/fixed-asset-results/{zip_filename}'
        S3_CLIENT.upload_file(f'/tmp/{zip_filename}', SMART_EYE_BUCKET, s3_key)
        
        presigned_url = S3_CLIENT.generate_presigned_url(
            ClientMethod='get_object',
            Params={
                'Bucket': SMART_EYE_BUCKET,
                'Key': s3_key
            },
            ExpiresIn=3600 * 24  # 24 hours
        )
        
        clean_temp_files([filename, zip_filename])
        
        return {
            'type': 'FixedAsset',
            'records': len(fixed_assets),
            'files': [filename],
            's3_path': f'{SMART_EYE_BUCKET}/{s3_key}',
            'presigned_url': presigned_url
        }
        
    except Exception as e:
        logger.error(f"Error exporting fixed assets: {str(e)}")
        raise

@tracer.capture_method
def exportAcquisitionJournal(merchant_id, filters, sort_field, sort_direction, item_id_list, job_id, output_type, exportExcelFormat, file_extension):
    """Export Acquisition Journal data using Elasticsearch"""
    try:
        journals_es = getDataFromES(
            merchant_id,
            sort_field,
            sort_direction,
            filters,
            item_id_list,
            'acquisitionJournal'
        )
        
        if not journals_es:
            logger.info("No acquisition journal data found")
            return None
        
        # Extract _source from ES results
        journals = [item['_source'] for item in journals_es]
        
        # Get mapping configuration (updated path)
        mapping_config = getMappingConfig(merchant_id, 'fixedAsset', output_type)
        
        # Get export configuration
        export_config = mapping_config.get('exportAcquisitionJournal', {})
        
        # Define headers from mapping config
        headers = export_config.get('headers', [
            "Posting Date", "Document Date", "External Document No",
            "Document No.", "Account Type", "Account No.",
            "Depreciation Book Code", "FA Posting Type", "Description",
            "Amount", "Bal. Account Type", "Bal. Account No.",
            "Outlet Dim Code"
        ])
        
        # Process data using mapping
        rows = processAcquisitionJournalWithMapping(journals, export_config)
        
        # Create file
        current_date_time = datetime.strftime((datetime.now() + timedelta(hours=8)), '%Y-%m-%d_%H:%M:%S')
        filename = f'AcquisitionJournal_{current_date_time}.{file_extension}'
        
        # Clean old files
        clean_temp_files([filename])
        
        # UPDATED: Export logic based on format (following ExportGeneratedPOFINAL pattern)
        if exportExcelFormat:
            # Export as Excel
            exportToExcel(f'/tmp/{filename}', headers, rows, export_config)
        else:
            # Export as CSV
            write_csv_file(filename, headers, rows)
        
        # Create zip file
        zip_filename = f'AcquisitionJournal_{current_date_time}.zip'
        createZipFile(zip_filename, [filename])
        
        # UPDATED: S3 path (changed bucket and prefix)
        s3_key = f'export/fixed-asset-results/{zip_filename}'
        S3_CLIENT.upload_file(f'/tmp/{zip_filename}', SMART_EYE_BUCKET, s3_key)
        
        # Generate presigned URL
        presigned_url = S3_CLIENT.generate_presigned_url(
            ClientMethod='get_object',
            Params={
                'Bucket': SMART_EYE_BUCKET,
                'Key': s3_key
            },
            ExpiresIn=3600 * 24  # 24 hours
        )
        
        # Cleanup temp files
        clean_temp_files([filename, zip_filename])
        
        return {
            'type': 'AcquisitionJournal',
            'records': len(journals),
            'files': [filename],
            's3_path': f'{SMART_EYE_BUCKET}/{s3_key}',
            'presigned_url': presigned_url
        }
        
    except Exception as e:
        logger.error(f"Error exporting acquisition journal: {str(e)}")
        raise

@tracer.capture_method
def createZipFile(zip_filename, files):
    """Create zip file containing multiple files"""
    zip_filepath = f'/tmp/{zip_filename}'
    
    with ZipFile(zip_filepath, 'w') as zipf:
        for file in files:
            file_path = f'/tmp/{file}'
            zipf.write(file_path, file)
    
    logger.info(f"Created zip file: {zip_filename} containing {len(files)} files")    


@tracer.capture_method
def getMerchantConfiguration(merchantId):
    """
    Get merchant configuration once and return structured data
    (Following ExportGeneratedPOFINAL.py pattern)
    """
    try:
        response = MERCHANT_DDB_TABLE.get_item(Key={'merchantId': merchantId})
        merchant = response.get('Item', {})
        
        # Extract all necessary fields
        custom_logics = merchant.get('customLogics', {})
        fixed_asset_config = merchant.get('fixedAssetConfig', {})
        
        merchant_config = {
            'merchantId': merchantId,
            'customLogics': custom_logics,
            'mappingPaths': {
                'faClassMapping': merchant.get('faClassMapping'),
                'depreciationMapping': merchant.get('depreciationMapping'),
            },
            'promptPaths': fixed_asset_config.get('promptPaths', {})
        }
        
        return merchant_config
        
    except Exception as e:
        logger.error(f"Error fetching merchant configuration: {str(e)}")
        # Return default configuration
        return {
            'merchantId': merchantId,
            'customLogics': {
                'exportExcelFormat': False
            },
            'mappingPaths': {
                'faClassMapping': None,
                'depreciationMapping': None,
            },
            'promptPaths': {}
        }

@tracer.capture_method
def getMappingConfig(merchant_id, document_type, output_type):
    """Get mapping configuration from S3 (updated path)"""
    s3_key = f"mapping/{merchant_id}/fixedAsset/{output_type}.json"  # ← Changed path
    
    try:
        response = S3_CLIENT.list_objects_v2(
            Bucket=SMART_EYE_BUCKET,
            Prefix=s3_key
        )
        if 'Contents' in response:
            response = S3_CLIENT.get_object(Bucket=SMART_EYE_BUCKET, Key=s3_key)
            mapping_config = json.loads(response['Body'].read().decode('utf-8'))
        else:
            # Fallback to default mapping
            default_mapping_key = f"mapping/default/fixedAsset/{output_type}_default.json"
            response = S3_CLIENT.get_object(Bucket=SMART_EYE_BUCKET, Key=default_mapping_key)
            mapping_config = json.loads(response['Body'].read().decode('utf-8'))
        
        return mapping_config
    except Exception as e:
        logger.warning(f"Error loading mapping config: {str(e)}")
        # Return basic default configuration
        return getDefaultFixedAssetMapping()

@tracer.capture_method
def processFixedAssetsWithMapping(fixed_assets, mapping_config):
    """Process fixed assets with mapping configuration"""
    result_rows = []
    document_fields = mapping_config.get('document_fields', {})
    field_formats = mapping_config.get('header_formats', {})
    date_parsing_config = mapping_config.get('date_parsing', {})
    headers = mapping_config.get('headers', [])

    for asset in fixed_assets:
        row = []
        for field in headers:
            mapped_field = document_fields.get(field)
            value = asset.get(mapped_field, '') if mapped_field else ''
            
            # Apply formatting if configured
            format_pattern = field_formats.get(field)
            if format_pattern and value != '':
                value = applyFormat(value, format_pattern, date_parsing_config)
            
            row.append(value)
        
        result_rows.append(row)

    return result_rows

@tracer.capture_method
def processAcquisitionJournalWithMapping(journals, mapping_config):
    """Process acquisition journal with mapping configuration"""
    result_rows = []
    document_fields = mapping_config.get('document_fields', {})
    field_formats = mapping_config.get('header_formats', {})
    date_parsing_config = mapping_config.get('date_parsing', {})
    headers = mapping_config.get('headers', [])

    for journal in journals:
        row = []
        for field in headers:
            mapped_field = document_fields.get(field)
            value = journal.get(mapped_field, '') if mapped_field else ''
            
            # Apply formatting if configured
            format_pattern = field_formats.get(field)
            if format_pattern and value != '':
                value = applyFormat(value, format_pattern, date_parsing_config)
            
            row.append(value)
        
        result_rows.append(row)

    return result_rows

# ADDED: Excel export functions from ExportGeneratedPOFINAL.py

@tracer.capture_method
def exportToExcel(file_path, headers, data_rows, mapping_config):
    """Export data to Excel format using pandas (copied from ExportGeneratedPOFINAL.py)"""
    try:
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Get Excel-specific configuration from mapping
        excel_config = mapping_config.get('excel_config', {})
        
        # Basic Excel export
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Get worksheet name from config or use default
            sheet_name = excel_config.get('sheet_name', 'Export Data')
            
            # Write DataFrame to Excel
            df.to_excel(
                writer, 
                sheet_name=sheet_name,
                index=False,
                startrow=excel_config.get('start_row', 0),
                startcol=excel_config.get('start_col', 0)
            )
            
            # Apply Excel formatting if openpyxl is available
            try:
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply enhanced formatting if configured
                applyExcelFormatting(workbook, worksheet, headers, excel_config, len(data_rows))
                
            except ImportError:
                logger.info("openpyxl styling not available, using basic Excel export")
        
        logger.info(f"Successfully exported {len(data_rows)} rows to Excel file: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        # Fallback to CSV if Excel export fails
        logger.info("Falling back to CSV export due to Excel error")
        csv_file_path = file_path.replace('.xlsx', '.csv')
        with open(csv_file_path, 'w', encoding='utf-8', newline='') as csvFile:
            writer = csv.writer(csvFile)
            writer.writerow(headers)
            writer.writerows(data_rows)
        # Rename back to xlsx for consistency
        os.rename(csv_file_path, file_path)

@tracer.capture_method
def write_csv_file(filename, header, rows):
    """Write CSV file with given filename, header, and rows"""
    with open(f'/tmp/{filename}', 'w', encoding='utf-8-sig') as csvFile:
        writer = csv.writer(csvFile)
        writer.writerow(header)
        writer.writerows(rows)

@tracer.capture_method
def clean_temp_files(file_list):
    """Remove temporary files if they exist"""
    for file in file_list:
        file_path = f'/tmp/{file}'
        if os.path.exists(file_path):
            os.remove(file_path)

@tracer.capture_method
def createCombinedExport(export_results, merchant_id):
    """Create combined zip file for multiple exports"""
    current_date_time = datetime.strftime((datetime.now() + timedelta(hours=8)), '%Y-%m-%d-%H-%M-%S')
    combined_zip = f'FixedAsset_Combined_{current_date_time}.zip'
    
    # Download all exported files from S3
    temp_files = []
    for result in export_results:
        s3_path = result['s3_path']
        s3_key = s3_path.replace(f'{SMART_EYE_BUCKET}/', '')
        local_file = f"/tmp/{os.path.basename(s3_key)}"
        S3_CLIENT.download_file(SMART_EYE_BUCKET, s3_key, local_file)
        temp_files.append(local_file)
    
    # Create combined zip
    with ZipFile(f'/tmp/{combined_zip}', 'w') as zipf:
        for file in temp_files:
            zipf.write(file, os.path.basename(file))
    
    # Upload combined zip
    s3_key = f'export/combined/{combined_zip}'
    S3_CLIENT.upload_file(f'/tmp/{combined_zip}', SMART_EYE_BUCKET, s3_key)
    
    # Generate presigned URL
    presigned_url = S3_CLIENT.generate_presigned_url(
        ClientMethod='get_object',
        Params={
            'Bucket': SMART_EYE_BUCKET,
            'Key': s3_key
        },
        ExpiresIn=3600 * 24  # 24 hours
    )
    
    # Cleanup
    for file in temp_files:
        os.remove(file)
    os.remove(f'/tmp/{combined_zip}')
    
    return {
        's3_path': f'{SMART_EYE_BUCKET}/{s3_key}',
        'presigned_url': presigned_url
    }

@tracer.capture_method
def updateDownloadJobStatus(job_id, status, message, s3_object_path=None, object_presigned_url=None):
    """Update download job status in DynamoDB"""
    if not DOWNLOAD_JOB_DDB_TABLE:
        logger.warning("Download job table not configured")
        return
    
    now = datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    
    update_expression = 'SET #st=:st, #msg=:msg, #ua=:ua'
    expression_attribute_names = {
        '#st': 'status',
        '#msg': 'message',
        '#ua': 'updatedAt'
    }
    expression_attribute_values = {
        ':st': status,
        ':msg': message,
        ':ua': now
    }
    
    if object_presigned_url:
        update_expression += ', #opurl=:opurl'
        expression_attribute_names['#opurl'] = 'objectPresignedUrl'
        expression_attribute_values[':opurl'] = object_presigned_url
    
    if s3_object_path:
        update_expression += ', #s3b=:s3b, #s3op=:s3op'
        expression_attribute_names['#s3b'] = 's3Bucket'
        expression_attribute_names['#s3op'] = 's3ObjectPath'
        expression_attribute_values[':s3b'] = SMART_EYE_BUCKET
        expression_attribute_values[':s3op'] = s3_object_path
    
    DOWNLOAD_JOB_DDB_TABLE.update_item(
        Key={'downloadJobId': job_id},
        UpdateExpression=update_expression,
        ExpressionAttributeNames=expression_attribute_names,
        ExpressionAttributeValues=expression_attribute_values
    )

@tracer.capture_method
def getDefaultFixedAssetMapping():
    """Provide default mapping configuration"""
    return {
        "exportFixedAsset": {
            "headers": [
                "No.", "Description", "Responsible Employee", "FA Class Code",
                "FA Subclass Code", "FA Location Code", "Search Description",
                "Acquired", "Blocked", "FA Posting Group", "Outlet Dim Code",
                "Vendor No"
            ],
            "document_fields": {
                "No.": "faNumber",
                "Description": "description",
                "Responsible Employee": "responsibleEmployee",
                "FA Class Code": "faClassCode",
                "FA Subclass Code": "faSubclassCode",
                "FA Location Code": "faLocationCode",
                "Search Description": "searchDescription",
                "Acquired": "acquired",
                "Blocked": "blocked",
                "FA Posting Group": "faPostingGroup",
                "Outlet Dim Code": "outletDimCode",
                "Vendor No": "vendorNo"
            },
            "header_formats": {},
            "date_parsing": {
                "dayfirst": True,
                "yearfirst": False
            }
        },
        "exportAcquisitionJournal": {
            "headers": [
                "Posting Date", "Document Date", "External Document No",
                "Document No.", "Account Type", "Account No.",
                "Depreciation Book Code", "FA Posting Type", "Description",
                "Amount", "Bal. Account Type", "Bal. Account No.",
                "Outlet Dim Code"
            ],
            "document_fields": {
                "Posting Date": "postingDate",
                "Document Date": "documentDate",
                "External Document No": "externalDocumentNo",
                "Document No.": "documentNo",
                "Account Type": "accountType",
                "Account No.": "accountNo",
                "Depreciation Book Code": "depreciationBookCode",
                "FA Posting Type": "faPostingType",
                "Description": "description",
                "Amount": "amount",
                "Bal. Account Type": "balAccountType",
                "Bal. Account No.": "balAccountNo",
                "Outlet Dim Code": "outletDimCode"
            },
            "header_formats": {
                "Posting Date": "{date:DD/MM/YYYY}",
                "Document Date": "{date:DD/MM/YYYY}"
            },
            "date_parsing": {
                "dayfirst": True,
                "yearfirst": False
            }
        }
    }

@tracer.capture_method
def updateDownloadJobStatus(jobId, status, message, s3ObjectPath=None, objectPresignedURL=None):
    now = datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    DOWNLOAD_JOB_DDB_TABLE.update_item(
        Key={
            'downloadJobId': jobId
        },
        UpdateExpression='SET #st=:st, #msg=:msg, #ua=:ua, #opurl=:opurl, #s3b=:s3b, #s3op=:s3op',
        ExpressionAttributeNames={
            '#st': 'status',
            '#msg': 'message',
            '#ua': 'updatedAt',
            '#opurl': 'objectPresignedUrl',
            '#s3b': 's3Bucket',
            '#s3op': 's3ObjectPath'
        },
        ExpressionAttributeValues={
            ':st': status,
            ':msg': message,
            ':ua': now,
            ':opurl': objectPresignedURL,
            ':s3b': SMART_EYE_BUCKET,
            ':s3op': s3ObjectPath
        }
    )

@tracer.capture_method
def applyFormat(value, format_pattern, date_parsing_config=None):
    if value is None or value == '':
        return value
    
    pattern_match = re.match(r'^(.*?)\{(\w+):([^}]+)\}(.*?)$', format_pattern)
    if not pattern_match:
        return str(value)
    
    prefix, format_type, format_spec, suffix = pattern_match.groups()
    
    formatters = {
        'date': lambda v: formatDate(v, format_spec, date_parsing_config)
    }
    
    formatter = formatters.get(format_type, lambda v: str(v))
    formatted_value = formatter(value)
    
    return prefix + formatted_value + suffix

def formatDate(value, format_spec, date_parsing_config=None):
    """Format date with specification"""
    if not value:
        return ''
    
    date_config = date_parsing_config or {'dayfirst': True, 'yearfirst': False}
    
    strftime_format = (format_spec
                      .replace('YYYY', '%Y')
                      .replace('MM', '%m')
                      .replace('DD', '%d')
                      .replace('HH', '%H')
                      .replace('mm', '%M')
                      .replace('ss', '%S'))
    
    try:
        # Handle different date formats
        if isinstance(value, str):
            # Parse string date
            parsed = parser.parse(
                value,
                dayfirst=date_config.get('dayfirst', True),
                yearfirst=date_config.get('yearfirst', False)
            )
        else:
            # Already a datetime object
            parsed = value
        
        return parsed.strftime(strftime_format)
    except (parser.ParserError, ValueError, TypeError, AttributeError):
        return str(value)

@tracer.capture_method
def exportToExcel(file_path, headers, data_rows, mapping_config):
    """
    Export data to Excel format using pandas
    """
    try:
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Get Excel-specific configuration from mapping
        excel_config = mapping_config.get('excel_config', {})
        
        # Basic Excel export
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Get worksheet name from config or use default
            sheet_name = excel_config.get('sheet_name', 'Export Data')
            
            # Write DataFrame to Excel
            df.to_excel(
                writer, 
                sheet_name=sheet_name,
                index=False,
                startrow=excel_config.get('start_row', 0),
                startcol=excel_config.get('start_col', 0)
            )
            
            # Apply Excel formatting if openpyxl is available
            try:
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply enhanced formatting if configured
                applyExcelFormatting(workbook, worksheet, headers, excel_config, len(data_rows))
                
            except ImportError:
                logger.info("openpyxl styling not available, using basic Excel export")
        
        logger.info(f"Successfully exported {len(data_rows)} rows to Excel file: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        # Fallback to CSV if Excel export fails
        logger.info("Falling back to CSV export due to Excel error")
        csv_file_path = file_path.replace('.xlsx', '.csv')
        with open(csv_file_path, 'w', encoding='utf-8', newline='') as csvFile:
            writer = csv.writer(csvFile)
            writer.writerow(headers)
            writer.writerows(data_rows)
        # Rename back to xlsx for consistency
        os.rename(csv_file_path, file_path)

@tracer.capture_method
def applyExcelFormatting(workbook, worksheet, headers, excel_config, row_count):
    """
    Apply Excel formatting with error handling
    """
    try:
        # Apply header formatting
        header_format = excel_config.get('header_format', {})
        if header_format:
            applyHeaderFormatting(worksheet, headers, header_format)
        
        # Apply column formatting
        column_formats = excel_config.get('column_formats', {})
        if column_formats:
            applyColumnFormatting(worksheet, headers, column_formats, row_count)
        
        # Auto-adjust column widths
        if excel_config.get('auto_adjust_columns', True):
            autoAdjustColumnWidths(worksheet)
            
    except ImportError:
        logger.info("openpyxl styling not available")
    except Exception as e:
        logger.warning(f"Error applying Excel formatting: {str(e)}")

@tracer.capture_method
def applyHeaderFormatting(worksheet, headers, header_format):
    """Apply formatting to Excel headers"""
    try:
        # Create header style
        header_font = Font(
            bold=header_format.get('bold', True),
            color=header_format.get('font_color', '000000'),
            size=header_format.get('font_size', 11)
        )
        
        header_fill = PatternFill(
            start_color=header_format.get('background_color', 'E6E6FA'),
            end_color=header_format.get('background_color', 'E6E6FA'),
            fill_type='solid'
        )
        
        header_alignment = Alignment(
            horizontal=header_format.get('alignment', 'center'),
            vertical='center',
            wrap_text=header_format.get('wrap_text', True)
        )
        
        # Apply to header row
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
    except Exception as e:
        logger.warning(f"Error applying header formatting: {str(e)}")

@tracer.capture_method
def applyColumnFormatting(worksheet, headers, column_formats, row_count):
    """Apply formatting to specific columns"""
    try:
        for header, format_config in column_formats.items():
            if header in headers:
                col_num = headers.index(header) + 1
                
                # Apply formatting to entire column
                for row_num in range(2, row_count + 2):  # Skip header row
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Apply number format if specified
                    if 'number_format' in format_config:
                        cell.number_format = format_config['number_format']
                        
    except Exception as e:
        logger.warning(f"Error applying column formatting: {str(e)}")

@tracer.capture_method
def autoAdjustColumnWidths(worksheet):
    """Auto-adjust column widths based on content"""
    try:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        logger.warning(f"Error auto-adjusting column widths: {str(e)}")

@tracer.capture_method
def getDataFromES(merchantId, sortField, sortDirection, filters, selectedItems, resultType=None):
    filterConditionMap = {
        'eq': 'match_phrase',
        'match': 'match',
        'matchPhrase': 'match_phrase',
        'matchPhrasePrefix': 'match_phrase_prefix',
        'gt': 'gt',
        'gte': 'gte',
        'lt': 'lt',
        'lte': 'lte',
        'wildcard': 'wildcard',
        'regexp': 'regexp',
        'contains': 'wildcard' 
    }

    if resultType == 'fixedAsset':
        es_index = 'fixedasset'  
    elif resultType == 'acquisitionJournal':
        es_index = 'acquisitionjournal' 
    else:
        es_index = 'fixedasset'

    url = f'https://{ES_DOMAIN_ENDPOINT}/{es_index}/_doc/_search'

    # Add merchant ID filter
    if filters.get('and') and len(filters.get('and')) > 0:
        filters['and'].append({'merchantId': {'eq': merchantId}})
    else: 
        filters['and'] = [{'merchantId': {'eq': merchantId}}]
        
    query = {'bool': {'must': []}}
    
    for andCondition in filters.get('and', []):
        if andCondition.get('or') is None:
            if andCondition.get('and'):
                for subAndCondition in andCondition.get('and', []):
                    filterField, filterConditionAndValue = list(subAndCondition.items())[0]
                    filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                    if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                        query['bool']['must'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                    elif filterCondition == 'exists':
                         query['bool']['must'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                    else:
                        query['bool']['must'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
            else:        
                filterField, filterConditionAndValue = list(andCondition.items())[0]
                filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                    query['bool']['must'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                elif filterCondition == 'exists':
                    query['bool']['must'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                elif filterCondition == 'wildcard' or filterCondition == 'contains':
                    query['bool']['must'].append({
                        'bool': {
                            'should': [
                                {
                                    'wildcard': {
                                        filterField: {
                                            'value': filterValue.lower(),
                                            'case_insensitive': True,
                                            'rewrite': 'constant_score'
                                        }
                                    }
                                },
                                {
                                    'wildcard': {
                                        f'{filterField}.keyword': {
                                            'value': filterValue.lower(),
                                            'case_insensitive': True,
                                            'rewrite': 'constant_score'
                                        }
                                    }
                                }
                            ],
                            'minimum_should_match': 1
                        }
                    })
                else:
                    query['bool']['must'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
        else:
            orConditionQuery = {'bool': {'should': []}}
            for orCondition in andCondition.get('or', []):
                filterField, filterConditionAndValue = list(orCondition.items())[0]
                filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                    orConditionQuery['bool']['should'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                elif filterCondition == 'exists':
                    orConditionQuery['bool']['should'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                else:
                    orConditionQuery['bool']['should'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
            query['bool']['must'].append(orConditionQuery)
            
    payload = dict()
    payload['query'] = query
    payload['sort'] = {sortField: {'order': sortDirection}}
    payload['size'] = 10000

    if selectedItems:
        payload['query']['bool']['must'].append({'ids': {'values': selectedItems}})
    
    payloadES = json.dumps(payload)
    headers = {
        'Content-Type': "application/json",
        'User-Agent': "PostmanRuntime/7.20.1",
        'Accept': "application/json, text/plain, */*",
        'Cache-Control': "no-cache",
        'Postman-Token': "1ae2b03c-ac6c-45f4-9b37-4f95b9b0102c,b678f18f-3ebe-458e-b63b-6ced7b74851f",
        'Host': ES_DOMAIN_ENDPOINT,
        'Accept-Encoding': "gzip, deflate, br",
        'Connection': "keep-alive",
        'cache-control': "no-cache",
    }
    
    try:
        response = requests.request("GET", url, data=payloadES, headers=headers, auth=AWSAUTH)
        responseText = json.loads(response.text)
        
        if 'error' in responseText:
            logger.error(f"Elasticsearch error: {responseText.get('error')}")
            raise BadRequestException("Invalid query statement")
            
        totalResp = responseText.get('hits').get('total').get('value')
        currentTotalResp = len(responseText.get('hits').get('hits'))

        resultList = responseText.get('hits').get('hits')
        
        # Handle pagination for large result sets
        while totalResp > currentTotalResp:
            payload["from"] = str(currentTotalResp)
            payloadES = json.dumps(payload)
            response = requests.request("GET", url, data=payloadES, headers=headers, auth=AWSAUTH)
            responseText = json.loads(response.text)
            if 'error' in responseText:
                raise BadRequestException("Invalid query statement")
            currentTotalResp += len(responseText.get('hits').get('hits'))
            resultList += responseText.get('hits').get('hits')

        logger.info(f"Retrieved {len(resultList)} items from Elasticsearch")
        return resultList
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Error connecting to Elasticsearch: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Error querying Elasticsearch: {str(e)}")
        raise