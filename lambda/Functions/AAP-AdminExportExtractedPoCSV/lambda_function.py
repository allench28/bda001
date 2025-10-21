import json
import boto3
from boto3.dynamodb.conditions import Key, Attr
from datetime import datetime, timedelta
import dateutil
import os
import requests
from requests_aws4auth import AWS4Auth
import csv
from aws_lambda_powertools import Logger, Tracer
from custom_exceptions import BadRequestException, ResourceNotFoundException
from zipfile import ZipFile
import re
from dateutil import parser
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

ES_DOMAIN_ENDPOINT = os.environ.get('ES_DOMAIN_ENDPOINT')
S3_BUCKET = os.environ.get('SMART_EYE_BUCKET')
DOWNLOAD_JOB_TABLE = os.environ.get('DOWNLOAD_JOB_TABLE')
DOCUMENT_UPLOAD_TABLE = os.environ.get("DOCUMENT_UPLOAD_TABLE")
EXTRACTED_PO_LINE_ITEMS_TABLE = os.environ.get('EXTRACTED_PO_LINE_ITEMS_TABLE')
EXTRACTED_PO_TABLE = os.environ.get('EXTRACTED_PO_TABLE')
SMART_EYE_BUCKET = os.environ.get('SMART_EYE_BUCKET')
MERCHANT_TABLE = os.environ.get('MERCHANT_TABLE')
EXTRACTED_DOCUMENTS_TABLE = os.environ.get('EXTRACTED_DOCUMENTS_TABLE')
EXTRACTED_DOCUMENTS_LINE_ITEMS_TABLE = os.environ.get('EXTRACTED_DOCUMENTS_LINE_ITEMS_TABLE')

CREDENTIALS = boto3.Session().get_credentials()
S3_CLIENT = boto3.client(
    's3',
    region_name='ap-southeast-5',
    endpoint_url='https://s3.ap-southeast-5.amazonaws.com'
)
DDB_RESOURCE = boto3.resource('dynamodb')

ACCESS_KEY = CREDENTIALS.access_key
SECRET_ACCESS_KEY = CREDENTIALS.secret_key
AWSAUTH = AWS4Auth(ACCESS_KEY, SECRET_ACCESS_KEY, 'ap-southeast-5', 'es', session_token=CREDENTIALS.token)

DOWNLOAD_JOB_DDB_TABLE = DDB_RESOURCE.Table(DOWNLOAD_JOB_TABLE)
DOCUMENT_UPLOAD_DDB_TABLE = DDB_RESOURCE.Table(DOCUMENT_UPLOAD_TABLE)
EXTRACTED_PO_LINE_ITEMS_DDB_TABLE = DDB_RESOURCE.Table(EXTRACTED_PO_LINE_ITEMS_TABLE)
EXTRACTED_PO_DDB_TABLE = DDB_RESOURCE.Table(EXTRACTED_PO_TABLE)
MERCHANT_DDB_TABLE = DDB_RESOURCE.Table(MERCHANT_TABLE)
EXTRACTED_DOCUMENT_DDB_TABLE = DDB_RESOURCE.Table(EXTRACTED_DOCUMENTS_TABLE)
EXTRACTED_DOCUMENTS_LINE_ITEMS_DDB_TABLE = DDB_RESOURCE.Table(EXTRACTED_DOCUMENTS_TABLE)

logger = Logger()
tracer = Tracer()

"""
SAMPLE PAYLOAD
ExtractedPO:
{
   "jobId":"e1885cbe-dbda-467b-98bd-9bd9cfcfc661",
   "arguments":{
      "resultType": "generate_po_so",
      "module":"ExtractedPo",
      "sort":{
         "field":"createdAt",
         "direction":"asc"
      },
      "filter": {"and": []}
   },
   "merchantId":"4d98df53-e473-445a-84e9-2681f1e82206"
}

GeneratedPO:
{
    "filter": {
        "and": []
    },
    "module": "GeneratedPo",
    "resultType": "generate_po",
    "sort": {
        "direction": "asc",
        "field": "createdAt"
    }
}
"""


@tracer.capture_lambda_handler
def lambda_handler(event, context):
    try:
        arguments = event.get('arguments')
        merchant_id = event.get('merchantId')
        itemIdList = arguments.get('itemIdList')
        jobId = event.get('jobId')
        resultType = arguments.get('resultType')  # 'po' or 'pos'
        output_type = arguments.get('outputType', 'csv')
        
        # ADDED: Get merchant configuration
        merchant_config = getMerchantConfiguration(merchant_id)
        
        # ADDED: Check if merchant wants Excel format and override output_type
        custom_logics = merchant_config.get('customLogics', {})
        exportExcelFormat = custom_logics.get('exportExcelFormat', False)
        if exportExcelFormat:
            output_type = 'xlsx'  # Override to use xlsx mapping
            file_extension = 'xlsx'
        else:
            file_extension = 'csv'

        # Sorting
        sortField = 'createdAt'
        sortDirection = 'desc'
        if arguments.get('sort'):
            sortField = arguments['sort'].get('field', 'createdAt')
            sortDirection = arguments['sort'].get('direction', 'desc')
            if sortField in [
                'itemCode', 'description', 'extractedPoLineItemsId', 'merchantId',
                'extractedPoId', 'documentUploadId', 'createdBy', 'updatedBy',
                'poNumber', 'supplierCode', 'supplierName'
            ]:
                sortField += '.keyword'

        filters = arguments.get('filter', {})
        
        # Get data from Elasticsearch
        extractedPo = getDataFromES(merchant_id, sortField, sortDirection, filters, itemIdList, resultType)
        
        currentDateTime = datetime.strftime(datetime.now() + timedelta(hours=8), '%Y_%m_%d_%H_%M_%S')
        
        if resultType == 'generate_po_so':
            # Export both extracted PO and SO in same ZIP file 
            mapping_config = getMappingConfig(merchant_id, 'po', output_type)
            
            # Process extracted PO
            po_section_config = mapping_config.get('exportExtractedPO', {})
            po_header = po_section_config.get('headers', [
                "PO Number", "PO Date", "Customer Name", "Customer Code", "Currency",
                "Delivery Address", "Requested Delivery Date", "Payment Terms", "Confidence Score",
                "Updated Date", "Document Name", "PO Status", "Issue Description"
            ])
            poProcessedResultCSVRows = processExtractedPOWithMapping(extractedPo, po_section_config)
            
            # Process SO
            so_section_config = mapping_config.get('generatedSO', {})
            so_header = so_section_config.get('headers', [
                "SO Number", "Customer Code", "Customer Name", "SO Date", "Currency",
                "Total Amount", "Tax Amount", "Grand Total", "Billing Address",
                "Delivery Address", "Source File"
            ])
            soProcessedResultCSVRows = processSOWithMapping(extractedPo, so_section_config)
            
            # UPDATED: Dynamic file naming based on export format
            poFilename = f'ExtractedPO{currentDateTime}.{file_extension}'
            soFilename = f'GeneratedSO{currentDateTime}.{file_extension}'
            zipFile = f'ExtractedData{currentDateTime}.zip'
            
            # No data check
            if not poProcessedResultCSVRows and not soProcessedResultCSVRows:
                print("error: No data found")
                updateDownloadJobStatus(jobId, 'COMPLETED', 'No data found')
                return True

            # Clean old files
            clean_temp_files([poFilename, soFilename, zipFile])

            # UPDATED: Export logic based on format
            if exportExcelFormat:
                # Export as Excel
                exportToExcel(f'/tmp/{poFilename}', po_header, poProcessedResultCSVRows, po_section_config)
                exportToExcel(f'/tmp/{soFilename}', so_header, soProcessedResultCSVRows, so_section_config)
            else:
                # Export as CSV (original logic)
                write_csv_file(poFilename, po_header, poProcessedResultCSVRows)
                write_csv_file(soFilename, so_header, soProcessedResultCSVRows)

            # Create ZIP with both files
            with ZipFile(f'/tmp/{zipFile}', 'w') as zipf:
                zipf.write(f'/tmp/{poFilename}', poFilename)
                zipf.write(f'/tmp/{soFilename}', soFilename)

            # Upload to S3
            s3_key = f'export/extraction-results/{zipFile}'
            
        elif resultType == 'generate_po':
            # Export generated PO only
            mapping_config = getMappingConfig(merchant_id, 'generatedPO', output_type)
            section_config = mapping_config.get('exportGeneratedPO', {})
            
            # Default headers if not in mapping
            default_headers = [
                "Document No", "Buy From Vendor Number", "Order Date", "Type", 
                "No.", "Quantity", "Location Code", "Unit of Measure Code", 
                "Line Discount Amount", "Purchaser Code", "Vendor Inv No", 
                "Buyer Group", "Dim", "PO Expiry", "Vendor Order No", "Vendor Shipment No",
                "Comment", "General Comment", "Expected Receipt"
            ]
            header = section_config.get('headers', default_headers)
            
            # Process with both tables 
            processedResultCSVRows = processGeneratedPOWithMapping(extractedPo, section_config)
            
            # UPDATED: Dynamic file naming based on export format
            filename = f'GeneratedPO{currentDateTime}.{file_extension}'
            zipFile = f'GeneratedPO{currentDateTime}.zip'
            
            # No data check
            if not processedResultCSVRows:
                print("error: No data found")
                updateDownloadJobStatus(jobId, 'COMPLETED', 'No data found')
                return True

            # Clean old files
            clean_temp_files([filename, zipFile])

            # UPDATED: Export logic based on format
            if exportExcelFormat:
                # Export as Excel
                exportToExcel(f'/tmp/{filename}', header, processedResultCSVRows, section_config)
            else:
                # Export as CSV (original logic)
                write_csv_file(filename, header, processedResultCSVRows)

            # Create ZIP
            with ZipFile(f'/tmp/{zipFile}', 'w') as zipf:
                zipf.write(f'/tmp/{filename}', filename)

            # Upload to S3
            s3_key = f'export/generatedPO-results/{zipFile}'
            
        else:
            raise BadRequestException(f"Invalid resultType: {resultType}")

        S3_CLIENT.upload_file(f'/tmp/{zipFile}', S3_BUCKET, s3_key)

        # Presigned URL
        objectPresignedURL = S3_CLIENT.generate_presigned_url(
            ClientMethod='get_object',
            Params={'Bucket': S3_BUCKET, 'Key': s3_key}
        )
        
        # Cleanup
        files_to_clean = [zipFile]
        if resultType == 'generate_po_so':
            files_to_clean.extend([poFilename, soFilename])
        elif resultType == 'generate_po':
            files_to_clean.append(filename)
            
        clean_temp_files(files_to_clean)

        # Update job status
        updateDownloadJobStatus(jobId, 'COMPLETED', 'Job Completed', s3_key, objectPresignedURL)
        return True

    except (BadRequestException, ResourceNotFoundException) as ex:
        print("error:", str(ex))
        return {'status': False, 'message': str(ex)}

    except Exception as ex:
        updateDownloadJobStatus(jobId, 'FAILED', str(ex))
        tracer.put_annotation("lambda_error", "true")
        tracer.put_annotation("lambda_name", context.function_name)
        tracer.put_metadata("event", event)
        tracer.put_metadata("message", str(ex))
        logger.exception({"message": str(ex)})
        return {
            'status': False,
            'message': "The server encountered an unexpected condition that prevented it from fulfilling your request."
        }

@tracer.capture_method
def write_csv_file(filename, header, rows):
    """
    Write CSV file with given filename, header, and rows
    
    Args:
        filename (str): Name of the CSV file to create
        header (list): List of column headers
        rows (list): List of data rows
    """
    with open(f'/tmp/{filename}', 'w', encoding='utf-8-sig') as csvFile:
        writer = csv.writer(csvFile)
        writer.writerow(header)
        writer.writerows(rows)

@tracer.capture_method
def clean_temp_files(file_list):
    """
    Remove temporary files if they exist
    
    Args:
        file_list (list): List of filenames to remove
    """
    for file in file_list:
        file_path = f'/tmp/{file}'
        if os.path.exists(file_path):
            os.remove(file_path)


@tracer.capture_method
def getDataFromES(merchantId, sortField, sortDirection, filters, selectedItems, resultType=None):
    filterConditionMap = {
        'eq': 'match_phrase',
        'match': 'match',
        'matchPhrase': 'match_phrase',
        'matchPhrasePrefix': 'match_phrase_prefix',
        'gt': 'gt',
        'gte': 'gte',
        'lt': 'lt',
        'lte': 'lte',
        'wildcard': 'wildcard',
        'regexp': 'regexp'
    }

    url = f'https://{ES_DOMAIN_ENDPOINT}/extractedpo/_doc/_search'

    # Add merchant ID filter
    if filters.get('and') and len(filters.get('and')) > 0:
        filters['and'].append({'merchantId': {'eq': merchantId}})
    else: 
        filters['and'] = [{'merchantId': {'eq': merchantId}}]
        
    query = {'bool': {'must': []}}
    for andCondition in filters.get('and', []):
        if andCondition.get('or') is None:
            if andCondition.get('and'):
                for subAndCondition in andCondition.get('and', []):
                    filterField, filterConditionAndValue = list(subAndCondition.items())[0]
                    filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                    if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                        query['bool']['must'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                    elif filterCondition == 'exists':
                         query['bool']['must'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                    else:
                        query['bool']['must'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
            else:        
                filterField, filterConditionAndValue = list(andCondition.items())[0]
                filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                    query['bool']['must'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                elif filterCondition == 'exists':
                    query['bool']['must'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                elif filterCondition == 'wildcard':
                    query['bool']['must'].append({
                        'bool': {
                            'should': [
                                {
                                    'wildcard': {
                                        filterField: {
                                            'value': filterValue.lower(),
                                            'case_insensitive': True,
                                            'rewrite': 'constant_score'
                                        }
                                    }
                                },
                                {
                                    'wildcard': {
                                        f'{filterField}.keyword': {
                                            'value': filterValue.lower(),
                                            'case_insensitive': True,
                                            'rewrite': 'constant_score'
                                        }
                                    }
                                }
                            ],
                            'minimum_should_match': 1
                        }
                    })
                else:
                    query['bool']['must'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
        else:
            orConditionQuery = {'bool': {'should': []}}
            for orCondition in andCondition.get('or', []):
                filterField, filterConditionAndValue = list(orCondition.items())[0]
                filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                    orConditionQuery['bool']['should'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                elif filterCondition == 'exists':
                    orConditionQuery['bool']['should'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                else:
                    orConditionQuery['bool']['should'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
            query['bool']['must'].append(orConditionQuery)
            
    payload = dict()
    payload['query'] = query
    payload['sort'] = {sortField: {'order': sortDirection}}
    payload['size'] = 10000

    if selectedItems:
        payload['query']['bool']['must'].append({'ids': {'values': selectedItems}})
    
    payloadES = json.dumps(payload)
    headers = {
        'Content-Type': "application/json",
        'User-Agent': "PostmanRuntime/7.20.1",
        'Accept': "application/json, text/plain, */*",
        'Cache-Control': "no-cache",
        'Postman-Token': "1ae2b03c-ac6c-45f4-9b37-4f95b9b0102c,b678f18f-3ebe-458e-b63b-6ced7b74851f",
        'Host': ES_DOMAIN_ENDPOINT,
        'Accept-Encoding': "gzip, deflate, br",
        'Connection': "keep-alive",
        'cache-control': "no-cache",
    }
    response = requests.request("GET", url, data=payloadES, headers=headers, auth=AWSAUTH)
    responseText = json.loads(response.text)
    
    if 'error' in responseText:
        raise BadRequestException("Invalid query statement")
    totalResp = responseText.get('hits').get('total').get('value')
    currentTotalResp = len(responseText.get('hits').get('hits'))

    resultList = responseText.get('hits').get('hits')
    while totalResp > currentTotalResp:
        payload["from"] = str(currentTotalResp)
        payloadES = json.dumps(payload)
        response = requests.request("GET", url, data=payloadES, headers=headers, auth=AWSAUTH)
        responseText = json.loads(response.text)
        if 'error' in responseText:
            raise BadRequestException("Invalid query statement")
        currentTotalResp += len(responseText.get('hits').get('hits'))
        resultList += responseText.get('hits').get('hits')

    return resultList

@tracer.capture_method
def getMappingConfig(merchant_id, document_type, output_type):
    s3_key = f"mapping/{merchant_id}/{document_type}/{output_type}.json"
    response = S3_CLIENT.list_objects_v2(
        Bucket=SMART_EYE_BUCKET,
        Prefix=s3_key
    )
    if 'Contents' in response:
        response = S3_CLIENT.get_object(
            Bucket=SMART_EYE_BUCKET,
            Key=s3_key
        )
        mapping_config = json.loads(response['Body'].read().decode('utf-8'))
    else:
        default_mapping_key = f"mapping/default/{document_type}/{output_type}_default.json"
        response = S3_CLIENT.get_object(
            Bucket=SMART_EYE_BUCKET,
            Key=default_mapping_key
        )
        mapping_config = json.loads(response['Body'].read().decode('utf-8'))
    return mapping_config

@tracer.capture_method
def processExtractedPOWithMapping(extracted_pos, mapping_config):
    result_rows = []
    po_fields = mapping_config.get('po_fields', {})
    field_formats = mapping_config.get('header_formats', {})
    date_parsing_config = mapping_config.get('date_parsing', {})
    headers = mapping_config.get('headers', [])

    for po in extracted_pos:
        document = po['_source']
        if document.get('documentStatus') != "Success":
            continue
            
        documentUploadId = document.get('documentUploadId')
        document_upload = getDocumentUpload(documentUploadId)
        if documentUploadId:
            document['fileName'] = document_upload.get('fileName')
        
        row = []
        for field in headers:
            mapped_field = po_fields.get(field)
            value = document.get(mapped_field, '') if mapped_field else ''
            
            # Apply formatting if configured
            format_pattern = field_formats.get(field)
            if format_pattern and value != '':
                value = applyFormat(value, format_pattern, date_parsing_config)
            
            row.append(value)
        
        result_rows.append(row)

    return result_rows

@tracer.capture_method
def processSOWithMapping(extracted_pos, mapping_config):
    result_rows = []
    so_fields = mapping_config.get('so_fields', {})
    field_formats = mapping_config.get('header_formats', {})
    date_parsing_config = mapping_config.get('date_parsing', {})
    headers = mapping_config.get('headers', [])

    for po in extracted_pos:
        document = po['_source']
        if document.get('documentStatus') != "Success":
            continue
        documentUploadId = document.get('documentUploadId')
        document_upload = getDocumentUpload(documentUploadId)
        
        # Create SO number
        so_number = f"SO-{document.get('poNumber')}"
        document['soNumber'] = so_number
        
        if documentUploadId:
            document['fileName'] = document_upload.get('fileName')
        
        row = []
        for field in headers:
            field_config = so_fields.get(field)
            
            if isinstance(field_config, dict):
                if field_config.get('derive_from'):
                    # Handle derived fields like SO Number
                    base_value = document.get(field_config['derive_from'], '')
                    if field_config.get('format'):
                        value = field_config['format'].format(value=base_value)
                    else:
                        value = base_value
                elif field_config.get('generate') == 'now':
                    # Generate current timestamp
                    value = (datetime.now() + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    value = ''
            else:
                # Simple field mapping
                value = document.get(field_config, '') if field_config else ''
            
            format_pattern = field_formats.get(field)
            if format_pattern and value != '':
                value = applyFormat(value, format_pattern, date_parsing_config)
            
            row.append(value)
        
        result_rows.append(row)

    return result_rows

@tracer.capture_method
def processGeneratedPOWithMapping(extracted_pos, mapping_config):
    result_rows = []
    document_fields = mapping_config.get('document_fields', {})
    line_item_fields = mapping_config.get('line_item_fields', {})
    field_formats = mapping_config.get('header_formats', {})
    date_parsing_config = mapping_config.get('date_parsing', {})
    headers = mapping_config.get('headers', [])

    # Always process POs with their line items 
    for po in extracted_pos:
        document = po['_source']
        merchant_id = document.get('merchantId')
        extracted_po_id = document.get('extractedPoId')  # Use extractedPoId instead of poNumber
        
        # Get line items for this PO from DynamoDB using extractedPoId
        line_items = getGeneratedPOLineItems(extracted_po_id, merchant_id)

        if not line_items:
            # No line items found - Skip
            continue
        else:
            # Line items found - create a row for each line item
            for line_item in line_items:
                row = createRowFromPOLineItem(
                    line_item, 
                    document,
                    headers, 
                    document_fields, 
                    line_item_fields, 
                    field_formats, 
                    date_parsing_config
                )
                result_rows.append(row)
    
    return result_rows
    
@tracer.capture_method
def getGeneratedPOLineItems(extracted_po_id, merchant_id):
    try:
        # Use query instead of scan with extractedPoId as partition key
        response = EXTRACTED_PO_LINE_ITEMS_DDB_TABLE.query(
            IndexName='gsi-extractedPoId',
            KeyConditionExpression=Key('extractedPoId').eq(extracted_po_id)
        )
        items = response.get('Items', [])
        
        # Handle pagination if needed
        while 'LastEvaluatedKey' in response:
            response = EXTRACTED_PO_LINE_ITEMS_DDB_TABLE.query(
                IndexName='gsi-extractedPoId',
                KeyConditionExpression=Key('extractedPoId').eq(extracted_po_id),
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response.get('Items', []))
        
        # Filter by merchant_id if needed (assuming it's not part of the key)
        filtered_items = [item for item in items if item.get('merchantId') == merchant_id]
        
        return filtered_items
    except Exception as e:
        logger.error(f"Error fetching PO line items: {str(e)}")
        return []

@tracer.capture_method
def createRowFromPOLineItem(line_item, document, headers, document_fields, line_item_fields, field_formats, date_parsing_config):
    row = []
    extracted_po_id = document.get('extractedPoId')
    merchant_id = document.get('merchantId')

    for field in headers:
        doc_mapped_field = document_fields.get(field)
        item_mapped_field = line_item_fields.get(field)
        value = None

        # Dynamically access fields from extracted document row if needed
        if doc_mapped_field:
            value = document.get(doc_mapped_field, '')
            # logger.info(f'doc_mapped_field: {value}')
        elif item_mapped_field:
            value = line_item.get(item_mapped_field, '')
            # logger.info(f'item_mapped_field: {value}')
        else:
            value = ''

        format_pattern = field_formats.get(field)
        if format_pattern and value != '':
            value = applyFormat(value, format_pattern, date_parsing_config)

        row.append(value)

    return row

@tracer.capture_method
def applyFormat(value, format_pattern, date_parsing_config=None):
    if value is None or value == '':
        return value
    
    pattern_match = re.match(r'^(.*?)\{(\w+):([^}]+)\}(.*?)$', format_pattern)
    if not pattern_match:
        return str(value)
    
    prefix, format_type, format_spec, suffix = pattern_match.groups()
    
    formatters = {
        'date': lambda v: formatDate(v, format_spec, date_parsing_config)
    }
    
    formatter = formatters.get(format_type, lambda v: str(v))
    formatted_value = formatter(value)
    
    return prefix + formatted_value + suffix

@tracer.capture_method
def formatDate(value, spec, date_parsing_config=None):
    if not value:
        return ''
    
    date_config = date_parsing_config or {}
    
    strftime_format = (spec
                      .replace('YYYY', '%Y')
                      .replace('MM', '%m')
                      .replace('DD', '%d')
                      .replace('HH', '%H')
                      .replace('mm', '%M')
                      .replace('ss', '%S'))
    
    try:
        parsed = parser.parse(
            str(value), 
            dayfirst=date_config.get('dayfirst', True),
            yearfirst=date_config.get('yearfirst', False)
        )
        
        if 'T' in str(value) or ':' in str(value):
            parsed += timedelta(hours=8)
        
        return parsed.strftime(strftime_format)
    except (parser.ParserError, ValueError, TypeError):
        return str(value)

@tracer.capture_method
def updateDownloadJobStatus(jobId, status, message, s3ObjectPath=None, objectPresignedURL=None):
    now = datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    DOWNLOAD_JOB_DDB_TABLE.update_item(
        Key={
            'downloadJobId': jobId
        },
        UpdateExpression='SET #st=:st, #msg=:msg, #ua=:ua, #opurl=:opurl, #s3b=:s3b, #s3op=:s3op',
        ExpressionAttributeNames={
            '#st': 'status',
            '#msg': 'message',
            '#ua': 'updatedAt',
            '#opurl': 'objectPresignedUrl',
            '#s3b': 's3Bucket',
            '#s3op': 's3ObjectPath'
        },
        ExpressionAttributeValues={
            ':st': status,
            ':msg': message,
            ':ua': now,
            ':opurl': objectPresignedURL,
            ':s3b': S3_BUCKET,
            ':s3op': s3ObjectPath
        }
    )

@tracer.capture_method
def getDocumentUpload(documentUploadId):
    response = DOCUMENT_UPLOAD_DDB_TABLE.get_item(
        Key={'documentUploadId': documentUploadId}
    ).get('Item', {})
    return response

@tracer.capture_method
def getMerchantConfiguration(merchantId):
    """
    Get merchant configuration once and return structured data
    """
    try:
        response = MERCHANT_DDB_TABLE.get_item(Key={'merchantId': merchantId})
        merchant = response.get('Item', {})
        
        # Extract all necessary fields
        custom_logics = merchant.get('customLogics', {})
        mappingPrompts = merchant.get('mappingPrompts', {})
        
        merchant_config = {
            'merchantId': merchantId,
            'customLogics': custom_logics,
            'mappingPaths': {
                'supplierMapping': merchant.get('supplierMapping'),
                'itemMapping': merchant.get('itemMapping'),
                'storeMapping': merchant.get('storeMapping')
            },
            'promptPaths': mappingPrompts
        }
        
        return merchant_config
        
    except Exception as e:
        logger.error(f"Error fetching merchant configuration: {str(e)}")
        # Return default configuration
        return {
            'merchantId': merchantId,

            'customLogics': {
                'overrideQuantityFromUom': False,
                'useCustomerRefAsPO': False,
                'invoiceToPO': False,
                'useStoreMapping': False,
                'enableExceptionFields': False,
                'exportLineItemLevelDoc': False,
                'exportExcelFormat': False
            },
            'mappingPaths': {
                'supplierMapping': None,
                'itemMapping': None,
                'storeMapping': None
            },
            'promptPaths': {
                'vendorMappingPrompt': None,
                'itemMappingPrompt': None,
                'storeMappingPrompt': None,
                'exceptionCheckingPrompt': None
            }
        }
    

@tracer.capture_method
def exportToExcel(file_path, headers, data_rows, mapping_config):
    """
    Export data to Excel format using pandas
    """
    try:
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Get Excel-specific configuration from mapping
        excel_config = mapping_config.get('excel_config', {})
        
        # Basic Excel export
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Get worksheet name from config or use default
            sheet_name = excel_config.get('sheet_name', 'Export Data')
            
            # Write DataFrame to Excel
            df.to_excel(
                writer, 
                sheet_name=sheet_name,
                index=False,
                startrow=excel_config.get('start_row', 0),
                startcol=excel_config.get('start_col', 0)
            )
            
            # Apply Excel formatting if openpyxl is available
            try:
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply enhanced formatting if configured
                applyExcelFormatting(workbook, worksheet, headers, excel_config, len(data_rows))
                
            except ImportError:
                logger.info("openpyxl styling not available, using basic Excel export")
        
        logger.info(f"Successfully exported {len(data_rows)} rows to Excel file: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        # Fallback to CSV if Excel export fails
        logger.info("Falling back to CSV export due to Excel error")
        csv_file_path = file_path.replace('.xlsx', '.csv')
        with open(csv_file_path, 'w', encoding='utf-8', newline='') as csvFile:
            writer = csv.writer(csvFile)
            writer.writerow(headers)
            writer.writerows(data_rows)
        # Rename back to xlsx for consistency
        os.rename(csv_file_path, file_path)

@tracer.capture_method
def applyExcelFormatting(workbook, worksheet, headers, excel_config, row_count):
    """
    Apply Excel formatting with error handling
    """
    try:
        # Apply header formatting
        header_format = excel_config.get('header_format', {})
        if header_format:
            applyHeaderFormatting(worksheet, headers, header_format)
        
        # Apply column formatting
        column_formats = excel_config.get('column_formats', {})
        if column_formats:
            applyColumnFormatting(worksheet, headers, column_formats, row_count)
        
        # Auto-adjust column widths
        if excel_config.get('auto_adjust_columns', True):
            autoAdjustColumnWidths(worksheet)
            
    except ImportError:
        logger.info("openpyxl styling not available")
    except Exception as e:
        logger.warning(f"Error applying Excel formatting: {str(e)}")

@tracer.capture_method
def applyHeaderFormatting(worksheet, headers, header_format):
    """Apply formatting to Excel headers"""
    try:
        # Create header style
        header_font = Font(
            bold=header_format.get('bold', True),
            color=header_format.get('font_color', '000000'),
            size=header_format.get('font_size', 11)
        )
        
        header_fill = PatternFill(
            start_color=header_format.get('background_color', 'E6E6FA'),
            end_color=header_format.get('background_color', 'E6E6FA'),
            fill_type='solid'
        )
        
        header_alignment = Alignment(
            horizontal=header_format.get('alignment', 'center'),
            vertical='center',
            wrap_text=header_format.get('wrap_text', True)
        )
        
        # Apply to header row
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
    except Exception as e:
        logger.warning(f"Error applying header formatting: {str(e)}")

@tracer.capture_method
def applyColumnFormatting(worksheet, headers, column_formats, row_count):
    """Apply formatting to specific columns"""
    try:
        for header, format_config in column_formats.items():
            if header in headers:
                col_num = headers.index(header) + 1
                
                # Apply formatting to entire column
                for row_num in range(2, row_count + 2):  # Skip header row
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Apply number format if specified
                    if 'number_format' in format_config:
                        cell.number_format = format_config['number_format']
                        
    except Exception as e:
        logger.warning(f"Error applying column formatting: {str(e)}")

@tracer.capture_method
def autoAdjustColumnWidths(worksheet):
    """Auto-adjust column widths based on content"""
    try:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        logger.warning(f"Error auto-adjusting column widths: {str(e)}")

