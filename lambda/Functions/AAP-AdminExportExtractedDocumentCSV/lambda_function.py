import json
import boto3
from boto3.dynamodb.conditions import Key, Attr
from datetime import datetime, timedelta
import dateutil
import os
import requests
from requests_aws4auth import AWS4Auth
import csv
from aws_lambda_powertools import Logger, Tracer
from custom_exceptions import BadRequestException, ResourceNotFoundException
from zipfile import ZipFile
import re
from dateutil import parser
import copy
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

ES_DOMAIN_ENDPOINT = os.environ.get('ES_DOMAIN_ENDPOINT')
S3_BUCKET = os.environ.get('SMART_EYE_BUCKET')
DOWNLOAD_JOB_TABLE = os.environ.get('DOWNLOAD_JOB_TABLE')
DOCUMENT_UPLOAD_TABLE = os.environ.get("DOCUMENT_UPLOAD_TABLE")
EXTRACTED_DOCUMENTS_LINE_ITEMS_TABLE = os.environ.get('EXTRACTED_DOCUMENTS_LINE_ITEMS_TABLE')
SMART_EYE_BUCKET = os.environ.get('SMART_EYE_BUCKET')
BR_MERCHANT_ID = os.environ.get('BR_MERCHANT_ID')
MERCHANT_TABLE = os.environ.get('MERCHANT_TABLE')

CREDENTIALS = boto3.Session().get_credentials()
S3_CLIENT = boto3.client(
    's3',
    region_name='ap-southeast-5',
    endpoint_url='https://s3.ap-southeast-5.amazonaws.com'
)
DDB_RESOURCE = boto3.resource('dynamodb')

ACCESS_KEY = CREDENTIALS.access_key
SECRET_ACCESS_KEY = CREDENTIALS.secret_key
AWSAUTH = AWS4Auth(ACCESS_KEY, SECRET_ACCESS_KEY, 'ap-southeast-5', 'es', session_token=CREDENTIALS.token)

DOWNLOAD_JOB_DDB_TABLE = DDB_RESOURCE.Table(DOWNLOAD_JOB_TABLE)
DOCUMENT_UPLOAD_DDB_TABLE = DDB_RESOURCE.Table(DOCUMENT_UPLOAD_TABLE)
EXTRACTED_DOCUMENTS_LINE_ITEMS_DDB_TABLE = DDB_RESOURCE.Table(EXTRACTED_DOCUMENTS_LINE_ITEMS_TABLE)
MERCHANT_DDB_TABLE = DDB_RESOURCE.Table(MERCHANT_TABLE)

ZERO_WIDTH_SPACE = '\u200B'  # Prevents Excel from reformatting values

logger = Logger()
tracer = Tracer()

@tracer.capture_lambda_handler
def lambda_handler(event, context):
    try:
        arguments = event.get('arguments')
        merchant_id = event.get('merchantId')
        itemIdList = arguments.get('itemIdList')
        jobId = event.get('jobId')
        resultType = arguments.get('resultType')
        output_type = arguments.get('outputType', 'csv')
        document_type = arguments.get('documentType', 'invoice')
        merchant_config = getMerchantConfiguration(merchant_id)

        # ADDED: Check if merchant wants Excel format and override output_type
        custom_logics = merchant_config.get('customLogics', {})
        exportExcelFormat = custom_logics.get('exportExcelFormat', False)
        if exportExcelFormat:
            output_type = 'xlsx'  # Override to use xlsx mapping
            file_extension = 'xlsx'
        else:
            file_extension = 'csv'

        sortField = 'createdAt'
        sortDirection = 'desc'
        if arguments.get('sort') is not None:
            sortField = arguments.get('sort').get('field')
            sortDirection = arguments.get('sort').get('direction')
            if sortField in [
                'itemCode', 
                'description', 
                'extractedDocumentsLineItemsId', 
                'merchantId',
                'extractedDocumentsId',
                'documentUploadId',
                'createdBy',
                'updatedBy'
                ]:
                sortField += '.keyword'

        filters = arguments.get('filter', {})
        
        # Get documents from Elasticsearch with all filters
        extractedDocuments = getDataFromES(merchant_id, sortField, sortDirection, filters, itemIdList, resultType)
        mapping_config = getMappingConfig(merchant_id, document_type, output_type)
        if merchant_id == BR_MERCHANT_ID:
            extractedDocuments = performSupplierMapping(merchant_id, extractedDocuments)
        
        if resultType == "invoices":
            section_config = mapping_config.get('exportExtractedDocuments', {})
        else:  # items
            section_config = mapping_config.get('exportExtractedLineItems', {})

        header = section_config.get('headers', [])
        
        if resultType == "invoices":
            processedResultCSVRows = processDocumentWithMapping(
                extractedDocuments, 
                section_config,
                "invoices",
                None,
                None,
            )
        else:
            processedResultCSVRows = processDocumentWithMapping(
                extractedDocuments, 
                section_config,
                "items",
                merchant_config,
                filters,
            )
        
        if not processedResultCSVRows:
            print("error: ", 'No data found')
            updateDownloadJobStatus(jobId, 'COMPLETED', 'No data found')
            return True
        
        currentDateTime = datetime.strftime((datetime.now()+timedelta(hours=8)), '%Y_%m_%d_%H_%M_%S')

        if resultType == "invoices":
            filename = f'ExtractedDocument{currentDateTime}.{file_extension}'
            zipFile = f'ExtractedDocument{currentDateTime}.zip'
        else: # line items
            filename = f'ExtractedDocumentLineItems{currentDateTime}.{file_extension}'
            zipFile = f'ExtractedDocumentLineItems{currentDateTime}.zip'

        if os.path.exists('/tmp/' + filename):
            os.remove('/tmp/' + filename)

        # ADDED: Export logic based on format
        if exportExcelFormat:
            # Export as Excel
            exportToExcel('/tmp/' + filename, header, processedResultCSVRows, section_config)
        else:
            # Export as CSV (original logic)
            with open('/tmp/' + filename, 'a', encoding='utf-8') as csvFile:
                writer = csv.writer(csvFile)
                writer.writerow(header)
                writer.writerows(processedResultCSVRows)

        with ZipFile('/tmp/' + zipFile, 'w') as zip:
            zip.write('/tmp/' + filename, filename)

        S3_CLIENT.upload_file('/tmp/' + zipFile, S3_BUCKET, 'export/extraction-results/'+zipFile)

        objectPresignedURL = S3_CLIENT.generate_presigned_url(
            ClientMethod='get_object',
            Params={
                'Bucket': S3_BUCKET,
                'Key': 'export/extraction-results/'+zipFile
            }
        )
        os.remove('/tmp/' + filename)
        os.remove('/tmp/' + zipFile)
        updateDownloadJobStatus(jobId, 'COMPLETED', 'Job Completed', S3_BUCKET+'export/extraction-results/'+zipFile, objectPresignedURL)

        return True
    
    except (BadRequestException, ResourceNotFoundException) as ex:
        logger.exception({"message": str(ex)})
        return {'status': False, 'message': str(ex)}
    
    except Exception as ex:
        updateDownloadJobStatus(jobId, 'FAILED', str(ex))
        tracer.put_annotation("lambda_error", "true")
        tracer.put_annotation("lambda_name", context.function_name)
        tracer.put_metadata("event", event)
        tracer.put_metadata("message", str(ex))
        logger.exception({"message": str(ex)})
        return {'status': False, 'message': "The server encountered an unexpected condition that prevented it from fulfilling your request."}

@tracer.capture_method
def getDataFromES(merchantId, sortField, sortDirection, filters, selectedItems, resultType=None):
    filterConditionMap = {
            'eq': 'match_phrase',
            'match': 'match',
            'matchPhrase': 'match_phrase',
            'matchPhrasePrefix': 'match_phrase_prefix',
            'gt': 'gt',
            'gte': 'gte',
            'lt': 'lt',
            'lte': 'lte',
            'wildcard': 'wildcard',
            'regexp': 'regexp'
        }

    url = f'https://{ES_DOMAIN_ENDPOINT}/extracteddocuments/_doc/_search'

    # Add merchant ID filter
    if filters.get('and') and len(filters.get('and')) > 0:
        filters['and'].append({'merchantId': {'eq': merchantId}})
    else: 
        filters['and'] = [{'merchantId': {'eq': merchantId}}]
        
    
    query = {'bool': {'must': []}}
    for andCondition in filters.get('and', []):
        if andCondition.get('or') is None:
            if andCondition.get('and'):
                for subAndCondition in andCondition.get('and', []):
                    filterField, filterConditionAndValue = list(subAndCondition.items())[0]
                    filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                    if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                        query['bool']['must'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                    elif filterCondition == 'exists':
                         query['bool']['must'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                    else:
                        query['bool']['must'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
            else:        
                filterField, filterConditionAndValue = list(andCondition.items())[0]
                filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                    query['bool']['must'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                elif filterCondition == 'exists':
                    query['bool']['must'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                elif filterCondition == 'wildcard':
                    query['bool']['must'].append({
                        'bool': {
                            'should': [
                                {
                                    'wildcard': {
                                        filterField: {
                                            'value': filterValue.lower(),
                                            'case_insensitive': True,
                                            'rewrite': 'constant_score'
                                        }
                                    }
                                },
                                {
                                    'wildcard': {
                                        f'{filterField}.keyword': {
                                            'value': filterValue.lower(),
                                            'case_insensitive': True,
                                            'rewrite': 'constant_score'
                                        }
                                    }
                                }
                            ],
                            'minimum_should_match': 1
                        }
                    })
                else:
                    query['bool']['must'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
        else:
            orConditionQuery = {'bool': {'should': []}}
            for orCondition in andCondition.get('or', []):
                filterField, filterConditionAndValue = list(orCondition.items())[0]
                filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                if filterCondition == 'gt' or filterCondition == 'gte' or filterCondition == 'lt' or filterCondition == 'lte':
                    orConditionQuery['bool']['should'].append({"range":{filterField: {filterConditionMap[filterCondition]: filterValue}}})
                elif filterCondition == 'exists':
                    orConditionQuery['bool']['should'].append({filterConditionMap[filterCondition]: {"field": filterField}})
                else:
                    orConditionQuery['bool']['should'].append({filterConditionMap[filterCondition]: {filterField: filterValue}})
            query['bool']['must'].append(orConditionQuery)
            

    payload = dict()
    payload['query'] = query
    payload['sort'] = {sortField: {'order': sortDirection}}
    payload['size'] = 10000

    if selectedItems:
        payload['query']['bool']['must'].append({'ids': {'values': selectedItems}})
    
    payloadES = json.dumps(payload)
    headers = {
            'Content-Type': "application/json",
            'User-Agent': "PostmanRuntime/7.20.1",
            'Accept': "application/json, text/plain, */*",
            'Cache-Control': "no-cache",
            'Postman-Token': "1ae2b03c-ac6c-45f4-9b37-4f95b9b0102c,b678f18f-3ebe-458e-b63b-6ced7b74851f",
            'Host': ES_DOMAIN_ENDPOINT,
            'Accept-Encoding': "gzip, deflate, br",
            'Connection': "keep-alive",
            'cache-control': "no-cache",
        }
    response = requests.request("GET", url, data=payloadES, headers=headers, auth=AWSAUTH)
    responseText = json.loads(response.text)
    
    if 'error' in responseText:
        raise BadRequestException("Invalid query statement")
    totalResp = responseText.get('hits').get('total').get('value')
    currentTotalResp = len(responseText.get('hits').get('hits'))

    resultList = responseText.get('hits').get('hits')
    while totalResp > currentTotalResp:
        payload["from"] = str(currentTotalResp)
        payloadES = json.dumps(payload)
        response = requests.request("GET", url, data=payloadES, headers=headers, auth=AWSAUTH)
        responseText = json.loads(response.text)
        if 'error' in responseText:
            raise BadRequestException("Invalid query statement")
        currentTotalResp += len(responseText.get('hits').get('hits'))
        resultList += responseText.get('hits').get('hits')

    return resultList

@tracer.capture_method
def updateDownloadJobStatus(jobId, status, message, s3ObjectPath=None, objectPresignedURL=None):
    now = datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    DOWNLOAD_JOB_DDB_TABLE.update_item(
        Key={
            'downloadJobId': jobId
        },
        UpdateExpression='SET #st=:st, #msg=:msg, #ua=:ua, #opurl=:opurl, #s3b=:s3b, #s3op=:s3op',
        ExpressionAttributeNames={
            '#st': 'status',
            '#msg': 'message',
            '#ua': 'updatedAt',
            '#opurl': 'objectPresignedUrl',
            '#s3b': 's3Bucket',
            '#s3op': 's3ObjectPath'
        },
        ExpressionAttributeValues={
            ':st': status,
            ':msg': message,
            ':ua': now,
            ':opurl': objectPresignedURL,
            ':s3b': S3_BUCKET,
            ':s3op': s3ObjectPath
        }
    )

@tracer.capture_method
def getDocumentUpload(documentUploadId):
    response = DOCUMENT_UPLOAD_DDB_TABLE.get_item(
        Key={'documentUploadId': documentUploadId}
    ).get('Item', {})
    return response

@tracer.capture_method
def getExtractedDocumentLineItems(extracted_document_id):
    extractedDocLineItemsResp = EXTRACTED_DOCUMENTS_LINE_ITEMS_DDB_TABLE.query(
        IndexName='gsi-extractedDocumentsId',
        KeyConditionExpression=Key('extractedDocumentsId').eq(extracted_document_id)
    ).get('Items', [])
    
    return extractedDocLineItemsResp

@tracer.capture_method
def getMappingConfig(merchant_id, document_type, output_type):
    s3_key = f"mapping/{merchant_id}/{document_type}/{output_type}.json"
    response = S3_CLIENT.list_objects_v2(
        Bucket=SMART_EYE_BUCKET,
        Prefix=s3_key
    )
    if 'Contents' in response:
        response = S3_CLIENT.get_object(
            Bucket=SMART_EYE_BUCKET,
            Key=s3_key
        )
        mapping_config = json.loads(response['Body'].read().decode('utf-8'))
    else:
        default_mapping_key = f"mapping/default/{document_type}/{output_type}_default.json"
        response = S3_CLIENT.get_object(
            Bucket=SMART_EYE_BUCKET,
            Key=default_mapping_key
        )
        mapping_config = json.loads(response['Body'].read().decode('utf-8'))
    return mapping_config

@tracer.capture_method
def processDocumentWithMapping(extracted_documents, mapping_config, result_type, merchant_config, filters, status_values=None):
    result_rows = []
    document_fields = mapping_config.get('document_fields', {})
    line_item_fields = mapping_config.get('line_item_fields', {})
    field_formats = mapping_config.get('header_formats', {})
    date_parsing_config = mapping_config.get('date_parsing', {})
    headers = mapping_config.get('headers', [])
    single_document_line = mapping_config.get('single_document_line', False)
    repeated_document_fields = mapping_config.get('repeated_document_fields', [])

    if result_type == "invoices":
        # Process invoices at the document level
        for doc in extracted_documents:
            row = processInvoiceDocument(
                doc,
                document_fields,
                headers,
                field_formats,
                date_parsing_config
            )
            result_rows.append(row)

    elif result_type == "items":
        # ADDED: Extract status values from filters for line item level filtering
        status_values = extractStatusValuesFromFilters(filters)
        
        for doc in extracted_documents:
            rows = processLineItemsDocument(
                doc, 
                document_fields,
                line_item_fields, 
                headers,
                field_formats,
                date_parsing_config,
                status_values,
                merchant_config,
                single_document_line,
                repeated_document_fields,
            )
            result_rows.extend(rows)
    
    return result_rows

@tracer.capture_method
def extractStatusValuesFromFilters(filters):
    """
    Extract status values from filters to determine what line item statuses to include
    Following the pattern from existing filter processing
    """
    status_values = []
    
    if not filters:
        return None
    
    # Process filters to find documentStatus conditions
    for andCondition in filters.get('and', []):
        if 'or' in andCondition:
            for orCondition in andCondition['or']:
                filterField, filterConditionAndValue = list(orCondition.items())[0]
                if filterField == 'documentStatus':
                    filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                    if filterCondition in ['eq', 'match_phrase']:
                        # Map document status to line item status
                        if filterValue.lower() in ['exceptions', 'exception']:
                            status_values.append('exceptions')
                        elif filterValue.lower() == 'success':
                            status_values.append('success')
        else:
            filterField, filterConditionAndValue = list(andCondition.items())[0]
            if filterField == 'documentStatus':
                filterCondition, filterValue = list(filterConditionAndValue.items())[0]
                if filterCondition in ['eq', 'match_phrase']:
                    # Map document status to line item status
                    if filterValue.lower() in ['exceptions', 'exception']:
                        status_values.append('exceptions')
                    elif filterValue.lower() == 'success':
                        status_values.append('success')
    
    return status_values if status_values else None

def processInvoiceDocument(doc, document_fields, headers, field_formats, date_parsing_config):
    """Process a single invoice document"""
    document = doc['_source']
    document_upload_id = document.get('documentUploadId')
    
    # Get document upload info to add fileName
    document_upload = getDocumentUpload(document_upload_id)
    if document_upload_id:
        document['fileName'] = document_upload.get('fileName')
    
    row = []
    
    for field in headers:
        doc_mapped_field = document_fields.get(field)
        
        value = None
        if doc_mapped_field:
            value = document.get(doc_mapped_field, '')
        else:
            value = ''
        
        format_pattern = field_formats.get(field)
        if format_pattern and value != '':
            value = applyFormat(value, format_pattern, date_parsing_config)
        
        row.append(value)
    
    return row

def processLineItemsDocument(doc, document_fields, line_item_fields, headers, field_formats, date_parsing_config, status_values, merchant_config, single_document_line=False, repeated_document_fields=None):
    """
    Process a single document and its line items with merchant-specific filtering logic
    """
    rows = []
    document = doc['_source']
    document_upload_id = document.get('documentUploadId')
    extracted_documents_id = document.get('extractedDocumentsId')
    
    document_upload = getDocumentUpload(document_upload_id)
    if document_upload_id:
        document['fileName'] = document_upload.get('fileName')
    
    line_items = getExtractedDocumentLineItems(extracted_documents_id)
    
    # ADDED: Apply merchant-specific filtering logic
    if merchant_config:
        custom_logics = merchant_config.get('customLogics', {})
        exportLineItemLevelDoc = custom_logics.get('exportLineItemLevelDoc', False)
        
        if exportLineItemLevelDoc and status_values:
            # For line item level export: filter by individual line item status
            logger.info(f"Applying line item level filtering for status: {status_values}")
            line_items = [
                item for item in line_items 
                if item.get('status', '').lower() in [status.lower() for status in status_values]
            ]
        elif status_values:
            # For invoice level export: if invoice matches status filter, include all line items
            logger.info(f"Applying invoice level filtering - including all line items for filtered invoices")
            # Don't filter line items here - include all from the filtered invoices
            pass
    elif status_values:
        # Fallback to original behavior if no merchant config
        line_items = [
            item for item in line_items 
            if item.get('status', '').lower() in [status.lower() for status in status_values]
        ]
    
    if not line_items:
        return rows
    
    repeated_document_fields = repeated_document_fields or []
    
    for idx, line_item in enumerate(line_items):
        row = createRowFromLineItem(
            line_item, 
            document,
            headers, 
            document_fields, 
            line_item_fields, 
            field_formats, 
            date_parsing_config,
            single_document_line,
            is_first_row=(idx == 0),
            repeated_document_fields=repeated_document_fields
        )
        rows.append(row)
    
    return rows

def createRowFromLineItem(line_item, document, headers, document_fields, line_item_fields, field_formats, date_parsing_config, single_document_line=False, is_first_row=True, repeated_document_fields=None):
    row = []
    repeated_document_fields = repeated_document_fields or []
    
    for field in headers:
        doc_mapped_field = document_fields.get(field)
        item_mapped_field = line_item_fields.get(field)
        
        value = None
        
        # Check if this is a document field
        if doc_mapped_field:
            # For single_document_line mode, check if we should show this field
            if single_document_line and not is_first_row:
                # Only show value if this field is in repeated_document_fields
                if field in repeated_document_fields:
                    value = document.get(doc_mapped_field, '')
                else:
                    value = ''
            else:
                # First row or not single_document_line mode - show all values
                value = document.get(doc_mapped_field, '')
        elif item_mapped_field:
            value = line_item.get(item_mapped_field, '')
        else:
            value = ''
        
        format_pattern = field_formats.get(field)
        if format_pattern and value != '':
            value = applyFormat(value, format_pattern, date_parsing_config)
        
        row.append(value)
    
    return row

def applyFormat(value, format_pattern, date_parsing_config=None):
    """
    Apply formatting based on a pattern string.
    
    Args:
        value: The value to format
        format_pattern: Pattern like "{date:YYYY-MM-DD}"
        date_parsing_config: Optional configuration for date parsing
        
    Returns:
        Formatted string value
    """
    if value is None or value == '':
        return value
    
    pattern_match = re.match(r'^(.*?)\{(\w+)(?::([^}]*))?\}(.*?)$', format_pattern)
    if not pattern_match:
        return str(value)
    
    prefix, format_type, format_spec, suffix = pattern_match.groups()
    
    formatters = {
        'date': lambda v: formatDate(v, format_spec, date_parsing_config),
        'text': lambda v: f"'{v}"
    }
    
    formatter = formatters.get(format_type, lambda v: str(v))
    formatted_value = formatter(value)
    
    return prefix + formatted_value + suffix

def formatDate(value, spec, date_parsing_config=None):
    """
    Format date with strftime specification
    Example in mapping json config file:

        "header_formats": {
            "uploaded at": "{date:YYYY-MM-DD HH:mm:ss}",
        },
        // date_parsing doesnt matter if input date format is already in ISO 8601 format (YYYY-MM-DD)
        "date_parsing": {
            "dayfirst": true,
            "yearfirst": false
        }
    """
    if not value:
        return ''
    
    date_config = date_parsing_config or {}
    
    strftime_format = (spec
                      .replace('YYYY', '%Y')
                      .replace('MM', '%m')
                      .replace('DD', '%d')
                      .replace('HH', '%H')
                      .replace('mm', '%M')
                      .replace('ss', '%S'))
    
    try:
        # depending on input date format, we set the dayfirst and yearfirst flags
        # day-first=True for input dates with EU style (DD/MM/YYYY)
        # dayfirst=False for input dates US style (MM/DD/YYYY)
        # yearfirst=True for when year in input dates is 2 digits, ie YY/MM/DD
        parsed = parser.parse(
            str(value), 
            dayfirst=date_config.get('dayfirst', True),
            yearfirst=date_config.get('yearfirst', False)
        )
        
        if 'T' in str(value) or ':' in str(value):
            parsed += timedelta(hours=8)
        
        # return parsed.strftime(strftime_format) + ZERO_WIDTH_SPACE
        return f" {parsed.strftime(strftime_format)}"
    except (parser.ParserError, ValueError, TypeError):
        return str(value)
    
@tracer.capture_method
def performSupplierMapping(merchantId, extractedDocuments, limit=10, nextToken=0):
    extractedDocumentsData = copy.deepcopy(extractedDocuments)
    for document in extractedDocumentsData:
        supplierCode = document.get('_source').get('supplierCode')
        supplierName = document.get('_source').get('supplierName')
        
        if not supplierCode and not supplierName:
            continue
        
        # Get supplier details from Elasticsearch
        suppliers, total = getSupplierFromES(merchantId, document.get('_source'), limit, nextToken)
        
        if suppliers:
            logger.info(f'SUPPLIER: {suppliers}')
            # If we found suppliers, update the document with the first match
            erpBranchName = suppliers[0].get('erpBranchName', '')
            document['erpSupplierName'] = erpBranchName


    return extractedDocumentsData


@tracer.capture_method
def getSupplierFromES(merchant_id, extractedDocument, limit, nextToken):
    url = f"https://{ES_DOMAIN_ENDPOINT}/supplier/_search"

    must_clauses = [
        {"match_phrase": {"merchantId": merchant_id}}
    ]
    should_clauses = []

    # Define which fields use which kind of search
    match_phrase_fields = ["supplierCode"]
    wildcard_fields = ["supplierName"]

    for field in match_phrase_fields:
        value = extractedDocument.get(field)
        if value and value is not None:
            should_clauses.append({"match_phrase": {field: value}})

    for field in wildcard_fields:
        input_value = extractedDocument.get(field)
        if input_value and input_value is not None:
            tokens = re.findall(r'\w+', input_value.upper())
            wildcard_value = '*' + '*'.join(tokens) + '*'
            should_clauses.append({
                "bool": {
                    "should": [
                        {
                            "wildcard": {
                                f"{field}.keyword": {
                                    "value": wildcard_value,
                                }
                            }
                        }
                    ],
                    "minimum_should_match": 1
                }
            })

    query = {
        "bool": {
            "must": must_clauses,
            "should": should_clauses,
            "minimum_should_match": 2 if should_clauses else 0
        }
    }

    payload = {
        "query": query,
        "size": limit,
        "from": nextToken,
        "track_total_hits": True
    }

    headers = {
        'Content-Type': "application/json",
        'User-Agent': "PostmanRuntime/7.20.1",
        'Accept': "application/json, text/plain, */*",
        'Cache-Control': "no-cache",
        'Postman-Token': "1ae2b03c-ac6c-45f4-9b37-4f95b9b0102c,b678f18f-3ebe-458e-b63b-6ced7b74851f",
        'Host': ES_DOMAIN_ENDPOINT,
        'Accept-Encoding': "gzip, deflate, br",
        'Connection': "keep-alive",
        'cache-control': "no-cache",
    }

    response = requests.request("GET", url, data=json.dumps(payload), headers=headers, auth=AWSAUTH)
    responseText = json.loads(response.text)
    if 'error' in responseText:
        return [], 0

    total = responseText.get('hits', {}).get('total', {}).get('value', 0)
    items = [hit.get('_source', {}) for hit in responseText.get('hits', {}).get('hits', [])]

    return items, total

@tracer.capture_method
def getMerchantConfiguration(merchantId):
    """
    Get merchant configuration once and return structured data
    """
    try:
        response = MERCHANT_DDB_TABLE.get_item(Key={'merchantId': merchantId})
        merchant = response.get('Item', {})
        
        # Extract all necessary fields
        custom_logics = merchant.get('customLogics', {})
        mappingPrompts = merchant.get('mappingPrompts', {})
        
        merchant_config = {
            'merchantId': merchantId,
            'customLogics': custom_logics,
            'mappingPaths': {
                'supplierMapping': merchant.get('supplierMapping'),
                'itemMapping': merchant.get('itemMapping'),
                'storeMapping': merchant.get('storeMapping')
            },
            'promptPaths': mappingPrompts
        }
        
        return merchant_config
        
    except Exception as e:
        logger.error(f"Error fetching merchant configuration: {str(e)}")
        # Return default configuration
        return {
            'merchantId': merchantId,

            'customLogics': {
                'overrideQuantityFromUom': False,
                'useCustomerRefAsPO': False,
                'invoiceToPO': False,
                'useStoreMapping': False,
                'enableExceptionFields': False,
                'exportLineItemLevelDoc': False,
                'exportExcelFormat': False
            },
            'mappingPaths': {
                'supplierMapping': None,
                'itemMapping': None,
                'storeMapping': None
            },
            'promptPaths': {
                'vendorMappingPrompt': None,
                'itemMappingPrompt': None,
                'storeMappingPrompt': None,
                'exceptionCheckingPrompt': None
            }
        }
    

@tracer.capture_method
def exportToExcel(file_path, headers, data_rows, mapping_config):
    """
    Export data to Excel format using pandas
    """
    try:
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Get Excel-specific configuration from mapping
        excel_config = mapping_config.get('excel_config', {})
        
        # Basic Excel export
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Get worksheet name from config or use default
            sheet_name = excel_config.get('sheet_name', 'Export Data')
            
            # Write DataFrame to Excel
            df.to_excel(
                writer, 
                sheet_name=sheet_name,
                index=False,
                startrow=excel_config.get('start_row', 0),
                startcol=excel_config.get('start_col', 0)
            )
            
            # Apply Excel formatting if openpyxl is available
            try:
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply enhanced formatting if configured
                applyExcelFormatting(workbook, worksheet, headers, excel_config, len(data_rows))
                
            except ImportError:
                logger.info("openpyxl styling not available, using basic Excel export")
        
        logger.info(f"Successfully exported {len(data_rows)} rows to Excel file: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        # Fallback to CSV if Excel export fails
        logger.info("Falling back to CSV export due to Excel error")
        csv_file_path = file_path.replace('.xlsx', '.csv')
        with open(csv_file_path, 'w', encoding='utf-8', newline='') as csvFile:
            writer = csv.writer(csvFile)
            writer.writerow(headers)
            writer.writerows(data_rows)
        # Rename back to xlsx for consistency
        os.rename(csv_file_path, file_path)

@tracer.capture_method
def applyExcelFormatting(workbook, worksheet, headers, excel_config, row_count):
    """
    Apply Excel formatting with error handling
    """
    try:
        # Apply header formatting
        header_format = excel_config.get('header_format', {})
        if header_format:
            applyHeaderFormatting(worksheet, headers, header_format)
        
        # Apply column formatting
        column_formats = excel_config.get('column_formats', {})
        if column_formats:
            applyColumnFormatting(worksheet, headers, column_formats, row_count)
        
        # Auto-adjust column widths
        if excel_config.get('auto_adjust_columns', True):
            autoAdjustColumnWidths(worksheet)
            
    except ImportError:
        logger.info("openpyxl styling not available")
    except Exception as e:
        logger.warning(f"Error applying Excel formatting: {str(e)}")

@tracer.capture_method
def applyHeaderFormatting(worksheet, headers, header_format):
    """Apply formatting to Excel headers"""
    try:
        # Create header style
        header_font = Font(
            bold=header_format.get('bold', True),
            color=header_format.get('font_color', '000000'),
            size=header_format.get('font_size', 11)
        )
        
        header_fill = PatternFill(
            start_color=header_format.get('background_color', 'E6E6FA'),
            end_color=header_format.get('background_color', 'E6E6FA'),
            fill_type='solid'
        )
        
        header_alignment = Alignment(
            horizontal=header_format.get('alignment', 'center'),
            vertical='center',
            wrap_text=header_format.get('wrap_text', True)
        )
        
        # Apply to header row
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
    except Exception as e:
        logger.warning(f"Error applying header formatting: {str(e)}")

@tracer.capture_method
def applyColumnFormatting(worksheet, headers, column_formats, row_count):
    """Apply formatting to specific columns"""
    try:
        for header, format_config in column_formats.items():
            if header in headers:
                col_num = headers.index(header) + 1
                
                # Apply formatting to entire column
                for row_num in range(2, row_count + 2):  # Skip header row
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Apply number format if specified
                    if 'number_format' in format_config:
                        cell.number_format = format_config['number_format']
                        
    except Exception as e:
        logger.warning(f"Error applying column formatting: {str(e)}")

@tracer.capture_method
def autoAdjustColumnWidths(worksheet):
    """Auto-adjust column widths based on content"""
    try:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        logger.warning(f"Error auto-adjusting column widths: {str(e)}")